"""
@author : bigV
"""
"""
# 自己定义标签，向后看，指示未来ramp发生与否
# 缩小数据尺度，W—KW
"""

import numpy as np
from matplotlib import pyplot as plt
from sklearn.neural_network import MLPRegressor    #### MLP 感知机####
from sklearn.tree import ExtraTreeRegressor        #### ExtraTree极端随机树回归####
from sklearn import tree                           #### 决策树回归####
from sklearn.ensemble import BaggingRegressor      #### Bagging回归####
from sklearn.ensemble import AdaBoostRegressor     #### Adaboost回归
from sklearn import linear_model                   #### 线性回归####
from sklearn import svm                            #### SVM回归####
from sklearn import ensemble                       #### Adaboost回归####  ####3.7GBRT回归####  ####3.5随机森林回归####
from sklearn import neighbors                      #### KNN回归####
from lstmRegressor import lstmRegressor            #### lstm回归
from sklearn.preprocessing import StandardScaler
import os
import pandas as pd
from evaluate_data import *
import openpyxl
import warnings
warnings.filterwarnings("ignore")

##构建数据
def create_data(data, train_num, ahead_num):
    dataX1, dataX2 = [], []
    dataY1, dataY2 = [], []

    for i in range(train_num - ahead_num):
        # print(i)
        a = data[i:(i + ahead_num), 0]
        dataX1.append(a)
    for j in range(train_num - ahead_num, len(data) - ahead_num):
        b = data[j:(j + ahead_num), 0]
        dataX2.append(b)

    dataY1 = data[ahead_num:train_num, 0]
    dataY2 = data[train_num:, 0]
    return np.array(dataX1), np.array(dataY1), np.array(dataX2), np.array(dataY2)

##加载数据
def load_data(filename, ahead_num,N1):

    # dataset = pd.read_csv(filename, header=None)
    # dataset=dataset.values
    # dataset=dataset.astype('float32')
    # DOY	PST	Global Horiz [W/m^2]	Dry Bulb Temp [deg C]	Avg Wind Speed @ 30ft [m/s]
    # Peak Wind Speed @ 30ft [m/s]	Airmass
    # label:ramp up(ramp rate>100)(1), ramp down(ramp rate<-100(2), no ramp(0)
    # label:发生ramp event时往前数5min
    # label:发生ramp event时往前数10min

    dataset = pd.read_csv(filename,encoding='gbk')
    dataset = pd.DataFrame(dataset)
    # print("dataset:",dataset)

    # print("dataset2:",dataset['Global Horiz [W/m^2]'])
    # print("dataset7:",dataset['label:ramp up(ramp rate>100)(1), ramp down(ramp rate<-100(2), no ramp(0)'])
    # print("dataset2:",dataset.iloc[:,2])
    # print("dataset7:",dataset.iloc[:,7])
    ##List
    dataPOV = dataset['Global Horiz [W/m^2]']
    #dataPOV = dataset.iloc[:,2]
    #dataPOV=dataPOV.values
    #dataPOV=dataPOV.astype('float32')
    dataPOV = np.array(dataPOV).reshape(-1,1)

    print("dataPOV:",dataPOV.shape)
    # print(dataPOV[100][0])
    #
    # # #数据W—KW
    # temPOV = []
    # for ikw in range(dataPOV.shape[0]):
    #     temPOV.append(dataPOV[ikw][0]/1000)
    # dataPOV = np.array(temPOV).reshape(-1,1)
    #
    # print("dataPOV:", dataPOV.shape)
    # print(dataPOV[100][0])

    # global minLen
    # if minLen == 1:
    #     # #1分钟标签
    #     # dataFLAG = dataset.iloc[:,7]
    #     dataFLAG = dataset['label:ramp up(ramp rate>100)(1), ramp down(ramp rate<-100(2), no ramp(0)']
    # elif minLen == 5:
    #     # 5分钟标签
    #     # dataFLAG = dataset.iloc[:, 8]
    #     dataFLAG = dataset['label:发生ramp event时往前数5min']
    # else:
    #     # 10分钟标签
    #     # dataFLAG = dataset.iloc[:, 9]
    #     dataFLAG = dataset['label:发生ramp event时往前数10min']

    #dataFLAG = list(dataFLAG)
    #print(dataFLAG)

    ##dataFrame
    # dataPOV = dataset.iloc[:,2]
    # dataFLAG = dataset.iloc[:,7]

    N2 = N1+500

    dataAll = dataPOV[:N2,:]

    #dataFll = dataFLAG[N1:N2]

    #归一化
    global scaler
    scaler = StandardScaler(copy=True,with_mean=True,with_std=True)
    dataAll = scaler.fit_transform(dataAll)

    trainX, trainY, testX, testY = create_data(dataAll,N1,ahead_num)
    print("trainX", trainX.shape)
    print("trainY", trainY.shape)
    return trainX, trainY, testX, testY   #dataFll

def pre_model(model,trainX,trainY,testX):
    model.fit(trainX,trainY)
    predict = model.predict(testX)
    return predict

# #向后看
def deal_flag(predictData):
    flag_temp=[]
    global minLen
    print("minLen:",minLen)
    for i in range(len(predictData)-1):
        if (predictData[i]-predictData[i+1]) > (0.1*minLen):
            flag_temp.append(2)   # 下降趋势
        elif (predictData[i]-predictData[i+1]) < (-0.1*minLen):
            flag_temp.append(1)   # 上升趋势
        else:
            flag_temp.append(0)
    return flag_temp

# normal
def deal_zero_accuracy(flag1,flag2):
    """
    :param flag1: 真实标
    :param flag2: 预测标
    :return: accuracy
    """
    if len(flag1) != len(flag2):
        flag1 = flag1[:-1]
    else:
        flag1 = flag1
    # 实际0 label的数目
    zeroCount = 0
    for j in range(len(flag1)):
        if flag1[j] == 0:
            zeroCount = zeroCount+1
    print("zeroCount:",zeroCount)

    # 预测0 label的数目
    zeroCountPre = 0
    for jj in range(len(flag1)):
        if flag2[jj] == 0:
            zeroCountPre = zeroCountPre + 1
    print("zeroCountPre:", zeroCountPre)

    # #预测0 label正确率
    # 预测正确的0 label数目
    rightZeroCount = 0
    for i in range(len(flag2)):
        if flag1[i] == flag2[i] and flag1[i] == 0:
            rightZeroCount = rightZeroCount + 1
    print("rightZeroCount:",rightZeroCount)
    accuracyZero = rightZeroCount / zeroCount

    dataZeroNum=[]
    dataZeroNum.append(zeroCount)
    dataZeroNum.append(zeroCountPre)
    dataZeroNum.append(rightZeroCount)
    return accuracyZero * 100, dataZeroNum

# positive
def deal_one_accuracy(flag1,flag2):

    if len(flag1) != len(flag2):
        flag1 = flag1[:-1]
    else:
        flag1 = flag1

    # 实际1 label的数目
    oneCount = 0
    for j in range(len(flag1)):
        if flag1[j] == 1:
            oneCount = oneCount + 1
    print("oneCount:",oneCount)

    # 预测1 label的数目
    oneCountPre = 0
    for jj in range(len(flag2)):
        if flag2[jj] == 1:
            oneCountPre = oneCountPre + 1
    print("oneCountPre:",oneCountPre)

    rightOneCount = 0
    # #预测1 label正确率
    for i in range(len(flag2)):
        if flag1[i] == flag2[i] and flag1[i] == 1:
            rightOneCount = rightOneCount + 1
    print("rightOneCount:",rightOneCount)
    if oneCount == 0:
        oneCount = 1
    accuracyOne = rightOneCount / oneCount

    dataOneNum=[]
    dataOneNum.append(oneCount)
    dataOneNum.append(oneCountPre)
    dataOneNum.append(rightOneCount)

    return accuracyOne * 100, dataOneNum

# negative
def deal_two_accuracy(flag1,flag2):

    if len(flag1) != len(flag2):
        flag1 = flag1[:-1]
    else:
        flag1 = flag1
    # 实际2 label的数目
    twoCount = 0
    for j in range(len(flag1)):
        if flag1[j] == 2:
            twoCount = twoCount + 1
    print("twoCount:",twoCount)

    # 预测2 label的数目
    twoCountPre = 0
    for j in range(len(flag2)):
        if flag2[j] == 2:
            twoCountPre = twoCountPre + 1
    print("twoCountPre:",twoCountPre)

    rightTwoCount = 0
    # #预测2 label正确的数目
    for i in range(len(flag2)):
        if flag1[i] == flag2[i] and flag1[i] == 2:
            rightTwoCount = rightTwoCount + 1
    print("rightTwoCount:",rightTwoCount)

    if twoCount == 0:
        twoCount = 1
    accuracyTwo = rightTwoCount / twoCount

    dataTwoNum=[]
    dataTwoNum.append(twoCount)
    dataTwoNum.append(twoCountPre)
    dataTwoNum.append(rightTwoCount)

    return accuracyTwo * 100, dataTwoNum

# #全部的预测准确率
def deal_accuracy(flag1,flag2):
    """
    :param flag1: 真实标
    :param flag2: 预测标
    :return:
    """
    if len(flag1) != len(flag2):
        flag1 = flag1[:-1]
    else:
        flag1 = flag1

    rightCount = 0
    for i in range(len(flag2)):
        if flag1[i] == flag2[i]:
            rightCount = rightCount+1
    accuracy = rightCount/len(flag2)
    return accuracy*100, rightCount

# 保存数据
def save_result(predict_values,dataY,flag_method,flag_true,model_name,save_position,rmse,mae,mape,N,accuracy_method,dataZeroNum1,dataOneNum1,dataTwoNum1,rightCount1):

    #flag_true = flag_true[1:]
    global minLen
    global season
    global ahead_num
    filename = "C:\\Users\\Administrator\\Desktop\\pov_multi\\testdata_timeSeries\\autumn\\ramp_forcast_2014_autumn_10min.csv"
    p1 = os.path.exists(filename)

    if p1:
        workbook = openpyxl.load_workbook(filename)
        sheet_name = str(N) + "point_"+str(ahead_num)
        #防止重复sheet
        #a_sheet = wb.get_sheet_by_name('Sheet1')
        #wb.sheetnames
        sheetNames = workbook.sheetnames
        #sheetNames = workbook.get_sheet_names() #老版
        if sheet_name in sheetNames:
            #wb[sheetname]
            #sheet1 = workbook.get_sheet_by_name(sheet_name) #老版
            sheet1 = workbook[sheet_name]
        else:
            sheet1 = workbook.create_sheet(sheet_name)
        #sheet1 = workbook.create_sheet(sheet_name)

        save_position = 1+(save_position - 1)*16

        sheet1.cell(row=save_position, column=1,value = model_name)
        sheet1.cell(row=save_position + 1, column=1, value='true_data')
        sheet1.cell(row=save_position + 2, column=1, value='predict_data')
        sheet1.cell(row=save_position + 4, column=1, value='real_flag')
        sheet1.cell(row=save_position + 5, column=1, value='predict_flag')

        ## 存储数据
        for i in range(len(dataY)):
            sheet1.cell(row=save_position+1, column=i+2, value=dataY[i].astype(float))
            sheet1.cell(row=save_position+2, column=i+2, value=predict_values[i].astype(float))
        for j in range(len(flag_method)):
                sheet1.cell(row=save_position + 4, column=j + 2, value=flag_true[j])
                sheet1.cell(row=save_position + 5, column=j + 2, value=flag_method[j])

        sheet1.cell(row=save_position+7, column=1, value='RMSE')
        sheet1.cell(row=save_position+7, column=2, value=rmse)

        sheet1.cell(row=save_position+8, column=1, value='MAPE')
        sheet1.cell(row=save_position+8, column=2, value=mape.astype(float))

        sheet1.cell(row=save_position+9, column=1, value='MAE')
        sheet1.cell(row=save_position+9, column=2, value=mae.astype(float))

        sheet1.cell(row=save_position + 10, column=1, value='accuracy')
        sheet1.cell(row=save_position + 10, column=2, value=accuracy_method[0])
        sheet1.cell(row=save_position + 10, column=4, value='accuracy0')
        sheet1.cell(row=save_position + 10, column=5, value=accuracy_method[1])
        sheet1.cell(row=save_position + 10, column=7, value='accuracy1')
        sheet1.cell(row=save_position + 10, column=8, value=accuracy_method[2])
        sheet1.cell(row=save_position + 10, column=10, value='accuracy2')
        sheet1.cell(row=save_position + 10, column=11, value=accuracy_method[3])

        # #count
        sheet1.cell(row=save_position + 12, column=1, value='zeroCount')
        sheet1.cell(row=save_position + 12, column=2, value=dataZeroNum1[0])
        sheet1.cell(row=save_position + 12, column=4, value='zeroCountPre')
        sheet1.cell(row=save_position + 12, column=5, value=dataZeroNum1[1])
        sheet1.cell(row=save_position + 12, column=7, value='rightZeroCount')
        sheet1.cell(row=save_position + 12, column=8, value=dataZeroNum1[2])

        sheet1.cell(row=save_position + 13, column=1, value='oneCount')
        sheet1.cell(row=save_position + 13, column=2, value=dataOneNum1[0])
        sheet1.cell(row=save_position + 13, column=4, value='oneCountPre')
        sheet1.cell(row=save_position + 13, column=5, value=dataOneNum1[1])
        sheet1.cell(row=save_position + 13, column=7, value='rightOneCount')
        sheet1.cell(row=save_position + 13, column=8, value=dataOneNum1[2])

        sheet1.cell(row=save_position + 14, column=1, value='twoCount')
        sheet1.cell(row=save_position + 14, column=2, value=dataTwoNum1[0])
        sheet1.cell(row=save_position + 14, column=4, value='twoCountPre')
        sheet1.cell(row=save_position + 14, column=5, value=dataTwoNum1[1])
        sheet1.cell(row=save_position + 14, column=7, value='rightTwoCount')
        sheet1.cell(row=save_position + 14, column=8, value=dataTwoNum1[2])

        sheet1.cell(row=save_position + 15, column=1, value='rightCount')
        sheet1.cell(row=save_position + 15, column=2, value=rightCount1)

        workbook.save(filename)

    else:
        workbook = openpyxl.Workbook()
        #sheet_name = str(N) + "min"
        sheet_name = str(N) + "point_" + str(ahead_num)
        # 防止重复sheet
        # a_sheet = wb.get_sheet_by_name('Sheet1')
        # wb.sheetnames
        sheetNames = workbook.sheetnames
        # sheetNames = workbook.get_sheet_names() #老版
        if sheet_name in sheetNames:
            # wb[sheetname]
            # sheet1 = workbook.get_sheet_by_name(sheet_name) #老版
            sheet1 = workbook[sheet_name]
        else:
            sheet1 = workbook.create_sheet(sheet_name)
        #sheet1 = workbook.create_sheet(sheet_name)

        save_position = 1 + (save_position - 1) * 16

        sheet1.cell(row=save_position, column=1, value=model_name)
        sheet1.cell(row=save_position + 1, column=1, value='true_data')
        sheet1.cell(row=save_position + 2, column=1, value='predict_data')
        sheet1.cell(row=save_position + 4, column=1, value='real_flag')
        sheet1.cell(row=save_position + 5, column=1, value='predict_flag')

        # 写入
        for i in range(len(dataY)):
            sheet1.cell(row=save_position + 1, column=i + 2, value=dataY[i].astype(float))
            sheet1.cell(row=save_position + 2, column=i + 2, value=predict_values[i].astype(float))

        for j in range(len(flag_method)):
            sheet1.cell(row=save_position + 4, column=j + 2, value=flag_true[j])
            sheet1.cell(row=save_position + 5, column=j + 2, value=flag_method[j])


        sheet1.cell(row=save_position + 7, column=1, value='RMSE')
        sheet1.cell(row=save_position + 7, column=2, value=rmse)

        sheet1.cell(row=save_position + 8, column=1, value='MAPE')
        sheet1.cell(row=save_position + 8, column=2, value=mape.astype(float))

        sheet1.cell(row=save_position + 9, column=1, value='MAE')
        sheet1.cell(row=save_position + 9, column=2, value=mae.astype(float))

        # flag正确率
        sheet1.cell(row=save_position + 10, column=1, value='accuracy')
        sheet1.cell(row=save_position + 10, column=2, value=accuracy_method[0])
        sheet1.cell(row=save_position + 10, column=4, value='accuracy0')
        sheet1.cell(row=save_position + 10, column=5, value=accuracy_method[1])
        sheet1.cell(row=save_position + 10, column=7, value='accuracy1')
        sheet1.cell(row=save_position + 10, column=8, value=accuracy_method[2])
        sheet1.cell(row=save_position + 10, column=10, value='accuracy2')
        sheet1.cell(row=save_position + 10, column=11, value=accuracy_method[3])

        # #count
        sheet1.cell(row=save_position + 12, column=1, value='zeroCount')
        sheet1.cell(row=save_position + 12, column=2, value=dataZeroNum1[0])
        sheet1.cell(row=save_position + 12, column=4, value='zeroCountPre')
        sheet1.cell(row=save_position + 12, column=5, value=dataZeroNum1[1])
        sheet1.cell(row=save_position + 12, column=7, value='rightZeroCount')
        sheet1.cell(row=save_position + 12, column=8, value=dataZeroNum1[2])

        sheet1.cell(row=save_position + 13, column=1, value='oneCount')
        sheet1.cell(row=save_position + 13, column=2, value=dataOneNum1[0])
        sheet1.cell(row=save_position + 13, column=4, value='oneCountPre')
        sheet1.cell(row=save_position + 13, column=5, value=dataOneNum1[1])
        sheet1.cell(row=save_position + 13, column=7, value='rightOneCount')
        sheet1.cell(row=save_position + 13, column=8, value=dataOneNum1[2])

        sheet1.cell(row=save_position + 14, column=1, value='twoCount')
        sheet1.cell(row=save_position + 14, column=2, value=dataTwoNum1[0])
        sheet1.cell(row=save_position + 14, column=4, value='twoCountPre')
        sheet1.cell(row=save_position + 14, column=5, value=dataTwoNum1[1])
        sheet1.cell(row=save_position + 14, column=7, value='rightTwoCount')
        sheet1.cell(row=save_position + 14, column=8, value=dataTwoNum1[2])

        sheet1.cell(row=save_position + 15, column=1, value='rightCount')
        sheet1.cell(row=save_position + 15, column=2, value=rightCount1)

        workbook.save(filename)

def main():


    #random_state = np.random.RandomState(7)
    np.random.RandomState(7)

    # lookback number
    global ahead_num
    ahead_num = 8
    # training number
    N1 = 2000
    global minLen
    minLen = 10

    global season
    season = 'autumn'
    # #多分钟
    filename =  "C:\\Users\\Administrator\\Desktop\\pov_multi\\testdata_timeSeries\\autumn\\ramp_forcast_2014_autumn_30min.csv"

    # # X是二维数组，Y是一维数组
    # x_train, y_train, x_test, y_test,ramp_flag1 = load_data(filename, ahead_num,N1)
    x_train, y_train, x_test, y_test = load_data(filename, ahead_num, N1)
    print("x_train:",x_train.shape)
    print("y_train:", y_train.shape)
    print("x_test:", x_test.shape)
    print("y_test:", y_test.shape)

    ##====================多模型  multi single model
    model_DecisionTreeRegressor = tree.DecisionTreeRegressor()     #决策树
    model_RandomForestRegressor = ensemble.RandomForestRegressor(n_estimators=50)  # 随机森林
    model_GradientBoostingRegressor = ensemble.GradientBoostingRegressor(n_estimators=50)  # GDBT
    model_LinearRegression = linear_model.LinearRegression()  # 线性回归
    model_SVR = svm.SVR()                                        # SVR回归
    model_KNeighborsRegressor = neighbors.KNeighborsRegressor()  # KNN回归
    #
    model_ExtraTreeRegressor = ExtraTreeRegressor()  # extra tree
    model_MLP = MLPRegressor(solver='lbfgs', hidden_layer_sizes=(20, 20, 20), random_state=2)  # MLP
    model_BaggingRegressor = BaggingRegressor()   # bggingRegressor
    model_AdaboostRegressor = AdaBoostRegressor() # adaboostRegressor

   # model_lstmModel = lstmRegressor()          # lstm


    ##预测=============predict
    predict_decideTree = pre_model(model_DecisionTreeRegressor,x_train, y_train, x_test)
    predict_randomForest = pre_model(model_RandomForestRegressor,x_train, y_train, x_test)
    predict_linear = pre_model(model_LinearRegression, x_train, y_train, x_test)
    predict_svr = pre_model(model_SVR, x_train, y_train, x_test)
    predict_kNeighbors = pre_model(model_KNeighborsRegressor, x_train, y_train, x_test)
    predict_gradientBoosting = pre_model(model_GradientBoostingRegressor, x_train, y_train, x_test)
    predict_extraTree = pre_model(model_ExtraTreeRegressor, x_train, y_train, x_test)
    predict_mlp = pre_model(model_MLP, x_train, y_train, x_test)

    predict_bagging = pre_model(model_BaggingRegressor, x_train, y_train, x_test)
    predict_adaboost = pre_model(model_AdaboostRegressor,x_train,y_train,x_test)

    #predict_lstm = pre_model(model_lstmModel,x_train,y_train,x_test)

    ##===============反归一化
    global scaler
    predict_decideTree = scaler.inverse_transform(predict_decideTree)
    predict_randomForest = scaler.inverse_transform(predict_randomForest)
    predict_linear = scaler.inverse_transform(predict_linear)
    predict_svr = scaler.inverse_transform(predict_svr)
    predict_kNeighbors = scaler.inverse_transform(predict_kNeighbors)
    predict_gradientBoosting = scaler.inverse_transform(predict_gradientBoosting)
    predict_extraTree = scaler.inverse_transform(predict_extraTree)
    predict_mlp = scaler.inverse_transform(predict_mlp)
    predict_bagging = scaler.inverse_transform(predict_bagging)
    predict_adaboost = scaler.inverse_transform(predict_adaboost)

   # predict_lstm = scaler.inverse_transform(predict_lstm)


    dataY = scaler.inverse_transform(y_test)
    #ramp_flag = deal_flag(dataY)
    # print("ramp_flag:",ramp_flag)
    # dataY = [dataY]
    # dataY = np.reshape(dataY, (-1, 1))
    #print("dataY:",dataY)
    #print("dataY:",dataY)

    ####===================ramp flag=====================#####
    # flag_decideTree = deal_flag(predict_decideTree)
    # accuracy_decideTree_zero, dataZeroNum1 = deal_zero_accuracy(ramp_flag,flag_decideTree)
    # accuracy_decideTree_one, dataOneNum1 = deal_one_accuracy(ramp_flag, flag_decideTree)
    # accuracy_decideTree_two, dataTwoNum1 = deal_two_accuracy(ramp_flag, flag_decideTree)
    # accuracy_decideTree, rightCount1 = deal_accuracy(ramp_flag,flag_decideTree)
    #
    # flag_randomForest = deal_flag(predict_randomForest)
    # accuracy_randomForest_zero, dataZeroNum2 = deal_zero_accuracy(ramp_flag,flag_randomForest)
    # accuracy_randomForest_one, dataOneNum2 = deal_one_accuracy(ramp_flag, flag_randomForest)
    # accuracy_randomForest_two, dataTwoNum2 = deal_two_accuracy(ramp_flag, flag_randomForest)
    # accuracy_randomForest, rightCount2 = deal_accuracy(ramp_flag, flag_randomForest)
    #
    # flag_linear = deal_flag(predict_linear)
    # accuracy_linear_zero, dataZeroNum3 = deal_zero_accuracy(ramp_flag,flag_linear)
    # accuracy_linear_one, dataOneNum3 = deal_one_accuracy(ramp_flag, flag_linear)
    # accuracy_linear_two, dataTwoNum3 = deal_two_accuracy(ramp_flag, flag_linear)
    # accuracy_linear, rightCount3 = deal_accuracy(ramp_flag, flag_linear)
    #
    # flag_svr = deal_flag(predict_svr)
    # accuracy_svr_zero, dataZeroNum4 = deal_zero_accuracy(ramp_flag,flag_svr)
    # accuracy_svr_one, dataOneNum4 = deal_one_accuracy(ramp_flag, flag_svr)
    # accuracy_svr_two, dataTwoNum4 = deal_two_accuracy(ramp_flag, flag_svr)
    # accuracy_svr, rightCount4 = deal_accuracy(ramp_flag, flag_svr)
    #
    # flag_kNeighbors = deal_flag(predict_kNeighbors)
    # accuracy_kNeighbors_zero, dataZeroNum5 = deal_zero_accuracy(ramp_flag,flag_kNeighbors)
    # accuracy_kNeighbors_one, dataOneNum5 = deal_one_accuracy(ramp_flag, flag_kNeighbors)
    # accuracy_kNeighbors_two, dataTwoNum5 = deal_two_accuracy(ramp_flag, flag_kNeighbors)
    # accuracy_kNeighbors, rightCount5 = deal_accuracy(ramp_flag, flag_kNeighbors)
    #
    # flag_mlp = deal_flag(predict_mlp)
    # accuracy_mlp_zero, dataZeroNum6 = deal_zero_accuracy(ramp_flag,flag_mlp)
    # accuracy_mlp_one, dataOneNum6 = deal_one_accuracy(ramp_flag, flag_mlp)
    # accuracy_mlp_two, dataTwoNum6 = deal_two_accuracy(ramp_flag, flag_mlp)
    # accuracy_mlp, rightCount6 = deal_accuracy(ramp_flag, flag_mlp)
    #
    # flag_gradientBoosting = deal_flag(predict_gradientBoosting)
    # accuracy_gradientBoosting_zero, dataZeroNum7 = deal_zero_accuracy(ramp_flag,flag_gradientBoosting)
    # accuracy_gradientBoosting_one, dataOneNum7 = deal_one_accuracy(ramp_flag, flag_gradientBoosting)
    # accuracy_gradientBoosting_two, dataTwoNum7 = deal_two_accuracy(ramp_flag, flag_gradientBoosting)
    # accuracy_gradientBoosting,rightCount7 = deal_accuracy(ramp_flag, flag_gradientBoosting)
    #
    # flag_extraTree = deal_flag(predict_extraTree)
    # accuracy_extraTree_zero, dataZeroNum8 = deal_zero_accuracy(ramp_flag,flag_extraTree)
    # accuracy_extraTree_one, dataOneNum8 = deal_one_accuracy(ramp_flag, flag_extraTree)
    # accuracy_extraTree_two, dataTwoNum8 = deal_two_accuracy(ramp_flag, flag_extraTree)
    # accuracy_extraTree, rightCount8 = deal_accuracy(ramp_flag, flag_extraTree)
    #
    # flag_bagging = deal_flag(predict_bagging)
    # accuracy_bagging_zero, dataZeroNum9 = deal_zero_accuracy(ramp_flag,flag_bagging)
    # accuracy_bagging_one, dataOneNum9 = deal_one_accuracy(ramp_flag, flag_bagging)
    # accuracy_bagging_two, dataTwoNum9 = deal_two_accuracy(ramp_flag, flag_bagging)
    # accuracy_bagging, rightCount9 = deal_accuracy(ramp_flag, flag_bagging)
    #
    # flag_adaboost = deal_flag(predict_adaboost)
    # accuracy_adaboost_zero, dataZeroNum10 = deal_zero_accuracy(ramp_flag, flag_adaboost)
    # accuracy_adaboost_one, dataOneNum10 = deal_one_accuracy(ramp_flag, flag_adaboost)
    # accuracy_adaboost_two, dataTwoNum10 = deal_two_accuracy(ramp_flag, flag_adaboost)
    # accuracy_adaboost,rightCount10 = deal_accuracy(ramp_flag, flag_adaboost)

   #  flag_lstm = deal_flag(predict_lstm)
   #  accuracy_lstm_zero, dataZeroNum13 = deal_zero_accuracy(ramp_flag, flag_lstm)
   #  accuracy_lstm_one, dataOneNum13 = deal_one_accuracy(ramp_flag, flag_lstm)
   #  accuracy_lstm_two, dataTwoNum13 = deal_two_accuracy(ramp_flag, flag_lstm)
   #  accuracy_lstm, rightCount13 = deal_accuracy(ramp_flag, flag_lstm)
   #
   #  ######=========================evaluate and save=========================#####
   #
   #  # #####=========================decideTree========================##########
    mae_decideTree = MAE1(dataY,predict_decideTree)
    rmse_decideTree = RMSE1(dataY,predict_decideTree)
    mape_decideTree = MAPE1(dataY,predict_decideTree)
    print("======================================================")
    print("rmse_decideTree:",rmse_decideTree)
    print("mape_decideTree:",mape_decideTree)
    print("mae_decideTree:",mae_decideTree)
   #  print("accuracy_decideTree:", accuracy_decideTree)
   #  print("accuracy_decideTree_zero:", accuracy_decideTree_zero)
   #  print("accuracy_decideTree_one:", accuracy_decideTree_one)
   #  print("accuracy_decideTree_two:", accuracy_decideTree_two)
   #  accuracyDecideTree=[]
   #  accuracyDecideTree.append(accuracy_decideTree)
   #  accuracyDecideTree.append(accuracy_decideTree_zero)
   #  accuracyDecideTree.append(accuracy_decideTree_one)
   #  accuracyDecideTree.append(accuracy_decideTree_two)
   #  #save_result(predict_decideTree, dataY,flag_decideTree,ramp_flag, "decideTree", 1, rmse_decideTree,
   #            #  mae_decideTree, mape_decideTree, N1, accuracyDecideTree,dataZeroNum1,dataOneNum1,dataTwoNum1,rightCount1)
   #
   #  #=======================random forest
    rmse_randomForest = RMSE1(dataY,predict_randomForest)
    mape_randomForest = MAPE1(dataY,predict_randomForest)
    mae_randomForest = MAE1(dataY,predict_randomForest)
    print("mae_randomForest:", mae_randomForest)
    print("rmse_randomForest:",rmse_randomForest)
    print("mape_randomForest:",mape_randomForest)
   #  print("accuracy_randomForest:", accuracy_randomForest)
   #  print("aaccuracy_randomForest_zero:", accuracy_randomForest_zero)
   #  print("accuracy_randomForest_one:", accuracy_randomForest_one)
   #  print("accuracy_randomForest_two:", accuracy_randomForest_two)
   #  accuracyRandomForest=[]
   #  accuracyRandomForest.append(accuracy_randomForest)
   #  accuracyRandomForest.append(accuracy_randomForest_zero)
   #  accuracyRandomForest.append(accuracy_randomForest_one)
   #  accuracyRandomForest.append(accuracy_randomForest_two)
   #  #save_result(predict_randomForest, dataY,flag_randomForest,ramp_flag, "random forest", 2, rmse_randomForest,
   #             # mae_randomForest, mape_randomForest, N1, accuracyRandomForest,dataZeroNum2,dataOneNum2,dataTwoNum2,rightCount2)
   #
   #  ######=====================linear
    rmse_linear = RMSE1(dataY,predict_linear)
    mape_linear = MAPE1(dataY,predict_linear)
    mae_linear = MAE1(dataY,predict_linear)
    print("mae_linear:", mae_linear)
    print("rmse_linear:", rmse_linear)
    print("mape_linear:", mape_linear)
   #  print("accuracy_linear", accuracy_linear)
   #  print("aaccuracy_linear_zero:", accuracy_linear_zero)
   #  print("accuracy_linear_one:", accuracy_linear_one)
   #  print("accuracy_linear_two:", accuracy_linear_two)
   #  accuracyLinear=[]
   #  accuracyLinear.append(accuracy_linear)
   #  accuracyLinear.append(accuracy_linear_zero)
   #  accuracyLinear.append(accuracy_linear_one)
   #  accuracyLinear.append(accuracy_linear_two)
   #  #save_result(predict_linear, dataY,flag_linear,ramp_flag, "linear", 3, rmse_linear,
   #             # mae_linear, mape_linear, N1, accuracyLinear,dataZeroNum3,dataOneNum3,dataTwoNum3,rightCount3)
   #
   #  ######=====================SVR best
    rmse_svr = RMSE1(dataY,predict_svr)
    mape_svr = MAPE1(dataY,predict_svr)
    mae_svr = MAE1(dataY,predict_svr)
    print("mae_svr:", mae_svr)
    print("rmse_svr:", rmse_svr)
    print("mape_svr:", mape_svr)
   #  print("accuracy_svr:", accuracy_svr)
   #  print("aaccuracy_svr_zero:", accuracy_svr_zero)
   #  print("accuracy_svr_one:", accuracy_svr_one)
   #  print("accuracy_svr_two:", accuracy_svr_two)
   #  accuracySVR=[]
   #  accuracySVR.append(accuracy_svr)
   #  accuracySVR.append(accuracy_svr_zero)
   #  accuracySVR.append(accuracy_svr_one)
   #  accuracySVR.append(accuracy_svr_two)
   # # save_result(predict_svr, dataY,flag_svr,ramp_flag, "svr", 4, rmse_svr,
   #             # mae_svr, mape_svr, N1,accuracySVR,dataZeroNum4,dataOneNum4,dataTwoNum4,rightCount4)
   #
   #  ######===========================KNN
    rmse_kNeighbors = RMSE1(dataY,predict_kNeighbors)
    mape_kNeighbors = MAPE1(dataY,predict_kNeighbors)
    mae_kNeighbors = MAE1(dataY,predict_kNeighbors)
    print("mae_kNeighbors:", mae_kNeighbors)
    print("rmse_kNeighbors:", rmse_kNeighbors)
    print("mape_kNeighbors:", mape_kNeighbors)
   #  print("accuracy_kNeighbors:", accuracy_kNeighbors)
   #  print("aaccuracy_kNeighbors_zero:", accuracy_kNeighbors_zero)
   #  print("accuracy_kNeighbors_one:", accuracy_kNeighbors_one)
   #  print("accuracy_kNeighbors_two:", accuracy_kNeighbors_two)
   #  accuracyKNN=[]
   #  accuracyKNN.append(accuracy_kNeighbors)
   #  accuracyKNN.append(accuracy_kNeighbors_zero)
   #  accuracyKNN.append(accuracy_kNeighbors_one)
   #  accuracyKNN.append(accuracy_kNeighbors_two)
   #  #save_result(predict_kNeighbors, dataY,flag_kNeighbors,ramp_flag, "KNN", 5, rmse_kNeighbors,
   #              #mae_kNeighbors, mape_kNeighbors, N1, accuracyKNN,dataZeroNum5,dataOneNum5,dataTwoNum5,rightCount5)
   #
    rmse_mlp = RMSE1(dataY,predict_mlp)
    mape_mlp = MAPE1(dataY,predict_mlp)
    mae_mlp = MAE1(dataY,predict_mlp)
    print("mae_mlp:", mae_mlp)
    print("rmse_mlp:", rmse_mlp)
    print("mape_mlp:", mape_mlp)
   #  print("accuracy_mlp:", accuracy_mlp)
   #  print("aaccuracy_mlp_zero:", accuracy_mlp_zero)
   #  print("accuracy_mlp_one:", accuracy_mlp_one)
   #  print("accuracy_mlp_two:", accuracy_mlp_two)
   #  accuracyMLP=[]
   #  accuracyMLP.append(accuracy_mlp)
   #  accuracyMLP.append(accuracy_mlp_zero)
   #  accuracyMLP.append(accuracy_mlp_one)
   #  accuracyMLP.append(accuracy_mlp_two)
   #  #save_result(predict_mlp, dataY,flag_mlp,ramp_flag, "mlp", 6, rmse_mlp,
   #              #mae_mlp, mape_mlp, N1, accuracyMLP,dataZeroNum6,dataOneNum6,dataTwoNum6,rightCount6)
   #
    rmse_gradientBoosting = RMSE1(dataY,predict_gradientBoosting)
    mape_gradientBoosting = MAPE1(dataY,predict_gradientBoosting)
    mae_gradientBoosting = MAE1(dataY,predict_gradientBoosting)
    print("mae_gradientBoosting:", mae_gradientBoosting)
    print("rmse_gradientBoosting:", rmse_gradientBoosting)
    print("mape_gradientBoosting:", mape_gradientBoosting)
   #  print("accuracy_gradientBoosting:", accuracy_gradientBoosting)
   #  print("aaccuracy_gradientBoosting_zero:", accuracy_gradientBoosting_zero)
   #  print("accuracy_gradientBoosting_one:", accuracy_gradientBoosting_one)
   #  print("accuracy_gradientBoosting_two:", accuracy_gradientBoosting_two)
   #  accuracyGradientBoosting=[]
   #  accuracyGradientBoosting.append(accuracy_gradientBoosting)
   #  accuracyGradientBoosting.append(accuracy_gradientBoosting_zero)
   #  accuracyGradientBoosting.append(accuracy_gradientBoosting_one)
   #  accuracyGradientBoosting.append(accuracy_gradientBoosting_two)
   #  #save_result(predict_gradientBoosting, dataY,flag_gradientBoosting,ramp_flag, "gradientBoosting", 7, rmse_gradientBoosting,
   #            #  mae_gradientBoosting, mape_gradientBoosting, N1, accuracyGradientBoosting,dataZeroNum7,dataOneNum7,dataTwoNum7,rightCount7)
   #
    rmse_extraTree = RMSE1(dataY,predict_extraTree)
    mape_extraTree = MAPE1(dataY,predict_extraTree)
    mae_extraTree = MAE1(dataY,predict_extraTree)
    print("mae_extraTree:", mae_extraTree)
    print("rmse_extraTree:", rmse_extraTree)
    print("mape_extraTree:", mape_extraTree)
   #  print("accuracy_extraTree:", accuracy_extraTree)
   #  print("aaccuracy_extraTree_zero:", accuracy_extraTree_zero)
   #  print("accuracy_extraTree_one:", accuracy_extraTree_one)
   #  print("accuracy_extraTree_two:", accuracy_extraTree_two)
   #  accuracyExtraTree=[]
   #  accuracyExtraTree.append(accuracy_extraTree)
   #  accuracyExtraTree.append(accuracy_extraTree_zero)
   #  accuracyExtraTree.append(accuracy_extraTree_one)
   #  accuracyExtraTree.append(accuracy_extraTree_two)
   #  #save_result(predict_extraTree, dataY,flag_extraTree,ramp_flag, "extraTree", 8, rmse_extraTree,
   #        #      mae_extraTree, mape_extraTree, N1, accuracyExtraTree,dataZeroNum8,dataOneNum8,dataTwoNum8,rightCount8)
   #
    rmse_bagging = RMSE1(dataY,predict_bagging)
    mape_bagging = MAPE1(dataY,predict_bagging)
    mae_bagging = MAE1(dataY,predict_bagging)
    print("mae_bagging:", mae_bagging)
    print("rmse_bagging:", rmse_bagging)
    print("mape_bagging:", mape_bagging)
   #  print("accuracy_bagging:", accuracy_bagging)
   #  print("aaccuracy_bagging_zero:", accuracy_bagging_zero)
   #  print("accuracy_bagging_one:", accuracy_bagging_one)
   #  print("accuracy_bagging_two:", accuracy_bagging_two)
   #  accuracyBagging=[]
   #  accuracyBagging.append(accuracy_bagging)
   #  accuracyBagging.append(accuracy_bagging_zero)
   #  accuracyBagging.append(accuracy_bagging_one)
   #  accuracyBagging.append(accuracy_bagging_two)
   #  #save_result(predict_bagging, dataY,flag_bagging,ramp_flag, "bagging", 9, rmse_bagging,
   #           #   mae_bagging, mape_bagging, N1, accuracyBagging,dataZeroNum9,dataOneNum9,dataTwoNum9,rightCount9)
   #
    rmse_adaboost = RMSE1(dataY,predict_adaboost)
    mape_adaboost = MAPE1(dataY,predict_adaboost)
    mae_adaboost = MAE1(dataY,predict_adaboost)
    print("mae_adaboost:", mae_adaboost)
    print("rmse_bagging:", rmse_adaboost)
    print("mape_adaboost:", mape_adaboost)
   #  print("accuracy_adaboost:", accuracy_adaboost)
   #  print("aaccuracy_adaboost_zero:", accuracy_adaboost_zero)
   #  print("accuracy_adaboost_one:", accuracy_adaboost_one)
   #  print("accuracy_adaboost_two:", accuracy_adaboost_two)
   #  accuracyAdaboost = []
   #  accuracyAdaboost.append(accuracy_adaboost)
   #  accuracyAdaboost.append(accuracy_adaboost_zero)
   #  accuracyAdaboost.append(accuracy_adaboost_one)
   #  accuracyAdaboost.append(accuracy_adaboost_two)
   # # save_result(predict_adaboost, dataY,flag_adaboost,ramp_flag, "adaboostRegression", 10, rmse_adaboost,
   #              #mae_adaboost, mape_adaboost, N1, accuracyAdaboost,dataZeroNum10,dataOneNum10,dataTwoNum10,rightCount10)


    # rmse_lstm = RMSE1(dataY,predict_lstm)
    # mape_lstm = MAPE1(dataY,predict_lstm)
    # mae_lstm = MAE1(dataY,predict_lstm)
    # print("mae_lstm:", mae_lstm)
    # print("rmse_lstm:", rmse_lstm)
    # print("mape_lstm:", mape_lstm)
    # print("accuracy_lstm", accuracy_lstm)
    # print("aaccuracy_lstm_zero:", accuracy_lstm_zero)
    # print("accuracy_lstm_one:", accuracy_lstm_one)
    # print("accuracy_lstm_two:", accuracy_lstm_two)
    # accuracyLstm = []
    # accuracyLstm.append(accuracy_lstm)
    # accuracyLstm.append(accuracy_lstm_zero)
    # accuracyLstm.append(accuracy_lstm_one)
    # accuracyLstm.append(accuracy_lstm_two)
    #save_result(predict_lstm, dataY,flag_lstm,ramp_flag, "lstm", 11, rmse_lstm,
                #mae_lstm, mape_lstm, N1, accuracyLstm,dataZeroNum13,dataOneNum13,dataTwoNum13,rightCount13)


#     ###===============画图===========================
#     plt.figure(1,figsize=(15, 4))
#
#     plt.plot(dataY[:], "k", label="true",linewidth=1)
#     # plt.plot(predict_decideTree, "k", label="decideTree")
#     # plt.plot(predict_randomForest, "r", label="randomForest",linewidth=1)
#     plt.plot(predict_lstm[:], "r", label="lstm",linewidth=1)
#     #plt.plot(predict_svr, "g", label="svr",linewidth=1)
#     # plt.plot(np.arange(len(y_test)), predict_linear, "y", label="linear")
#     # plt.plot(np.arange(len(y_test)), predict_mlp, "k", label="mlp")
#     # plt.plot(np.arange(len(y_test)), predict_gradientBoosting, "b", label="gdB")
#     plt.xlabel("min")
#     plt.ylabel("pov")
#     plt.title(str(N1)+"min")
#     plt.legend(loc='best')
# #    plt.savefig("resultdata_timeSeries\\"+season+"\\"+str(minLen)+"min"+"\\"+season+"_pov_"+str(N1)+"point.png", bbox_inches='tight')  # fig.savefig
#     plt.show()

if __name__=="__main__":

    main()



